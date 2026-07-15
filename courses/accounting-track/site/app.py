"""Local preview server for the Skarion Accounting Track course.

Run:  python app.py   (serves http://127.0.0.1:5057)
Reads course/module/lesson/assignment/resource/finding data from the SQLite
content store and renders the actual course files (docx/xlsx/pdf/md) inline.
"""
import base64
import io
import os
import sqlite3

import mammoth
import markdown as md
import openpyxl
from flask import Flask, abort, render_template_string, send_file, url_for

_FORMULA_CACHE = {}


def get_computed_values(path):
    """Recalculate a workbook with the `formulas` engine and return {SHEETNAME: {coord: value}}.

    openpyxl doesn't evaluate formulas; any file this project re-saved with openpyxl loses its
    cached formula results (data_only reads come back None). This recovers real numbers for the
    preview without touching the underlying .xlsx files. Cached in-process by (path, mtime).
    """
    mtime = os.path.getmtime(path)
    key = (path, mtime)
    if key in _FORMULA_CACHE:
        return _FORMULA_CACHE[key]
    result = {}
    try:
        import formulas
        xl = formulas.ExcelModel().loads(path).finish()
        sol = xl.calculate()
        fname = os.path.basename(path)
        prefix = f"'[{fname}]"
        for k, v in sol.items():
            if not isinstance(k, str) or not k.startswith(prefix):
                continue
            sheet_part, coord = k[len(prefix):].split("'!")
            try:
                arr = v.value
                val = arr[0][0]
            except Exception:
                continue
            result.setdefault(sheet_part.upper(), {})[coord] = val
    except Exception as e:
        result = {"__error__": str(e)}
    _FORMULA_CACHE[key] = result
    return result

TRACK_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DB_PATH = os.path.join(TRACK_ROOT, "db", "skarion_accounting_track.db")
MODERN_ROOT = os.path.abspath(os.path.join(TRACK_ROOT, "..", "..", "claude-made-course"))
MODERN_MODULES = os.path.join(MODERN_ROOT, "modules")

app = Flask(__name__)


def db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


STATUS_COLOR = {
    "ready": "#1e8e5a",
    "needs_revision": "#b8860b",
    "needs_redo": "#c23b3b",
    "draft": "#6b7280",
}
SEV_COLOR = {
    "blocker": "#c23b3b",
    "error": "#d9822b",
    "improvement": "#b8860b",
    "nice_to_have": "#4b7bd6",
}

BASE = """
<!doctype html><html><head><meta charset="utf-8">
<title>{{ title }} — Skarion Accounting Track</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
  :root { color-scheme: light dark; }
  body { font-family: -apple-system, Segoe UI, Roboto, sans-serif; max-width: 980px; margin: 0 auto; padding: 24px 20px 80px;
         background: #0f1115; color: #e6e8eb; line-height: 1.55; }
  a { color: #7db4ff; }
  header.top { display:flex; align-items:center; justify-content:space-between; margin-bottom: 28px; flex-wrap:wrap; gap:10px;}
  header.top .brand { font-weight: 700; font-size: 1.15rem; }
  nav.crumbs { font-size: .85rem; color:#9aa4b2; margin-bottom: 18px; }
  nav.crumbs a { color:#9aa4b2; text-decoration:none; }
  nav.crumbs a:hover { text-decoration:underline; }
  h1 { font-size: 1.6rem; margin-bottom: 4px; }
  h2 { font-size: 1.2rem; margin-top: 2.2rem; border-bottom: 1px solid #262b33; padding-bottom: 6px; }
  .tagline { color:#9aa4b2; margin-top:0; }
  .grid { display:grid; grid-template-columns: repeat(auto-fill,minmax(280px,1fr)); gap:14px; margin: 18px 0; }
  .card { background:#171a21; border:1px solid #262b33; border-radius:10px; padding:16px 18px; text-decoration:none; color:inherit; display:block; transition:border-color .15s; }
  .card:hover { border-color:#3d4552; }
  .card h3 { margin:0 0 6px; font-size:1.02rem; }
  .card p { margin:0; color:#9aa4b2; font-size:.88rem; }
  .badge { display:inline-block; font-size:.72rem; font-weight:600; padding:2px 9px; border-radius:99px; color:#0f1115; margin-bottom:8px; }
  .pill { display:inline-block; font-size:.72rem; padding:2px 8px; border-radius:6px; background:#20242c; color:#9aa4b2; margin-right:6px; }
  table { border-collapse: collapse; width: 100%; margin: 10px 0 20px; font-size:.9rem; }
  th, td { border: 1px solid #262b33; padding: 7px 10px; text-align: left; vertical-align:top; }
  th { background: #171a21; }
  tr:nth-child(even) td { background: #14171d; }
  .findings li { margin-bottom: 8px; }
  .sev { font-weight:700; text-transform:uppercase; font-size:.72rem; letter-spacing:.03em; }
  .status-open { color:#c23b3b; }
  .status-fixed_in_repo_copy { color:#1e8e5a; }
  .resource-row { display:flex; justify-content:space-between; align-items:center; padding:9px 0; border-bottom:1px solid #1c2027; }
  .resource-row .actions a { margin-left:10px; font-size:.85rem; }
  .doc-preview { background:#fff; color:#111; border-radius:8px; padding:28px 34px; margin-top:14px; max-width: 100%; overflow-x:auto; }
  .doc-preview table { font-size:.85rem; }
  .doc-preview img { max-width:100%; }
  .xlsx-sheet { margin-bottom:26px; }
  .xlsx-sheet h4 { margin-bottom:6px; }
  .tabbar { display:flex; gap:6px; flex-wrap:wrap; margin: 10px 0 14px; }
  .tabbar a { padding:4px 10px; border-radius:6px; background:#20242c; font-size:.82rem; text-decoration:none; color:#cbd2db; }
  .tabbar a.active { background:#3d68c9; color:#fff; }
  .banner { padding:10px 14px; border-radius:8px; margin: 10px 0 18px; font-size:.88rem; }
  .banner.danger { background:#3a1c1c; border:1px solid #6b2c2c; color:#f5c2c2; }
  .banner.warn { background:#3a301c; border:1px solid #6b5a2c; color:#f0dca6; }
  .muted { color:#9aa4b2; }
  code { background:#1c2027; padding:1px 5px; border-radius:4px; }
  .markdown-body h1, .markdown-body h2, .markdown-body h3 { border:none; }
  .markdown-body table { font-size:.88rem; }
  footer { margin-top:60px; color:#5b6472; font-size:.8rem; text-align:center; }
</style>
</head><body>
<header class="top">
  <div class="brand"><a href="{{ url_for('home') }}" style="text-decoration:none;color:inherit;">🧮 Skarion Accounting Track</a></div>
  <nav>
    <a href="{{ url_for('home') }}">Dashboard</a> ·
    <a href="{{ url_for('feedback_index') }}">Feedback for Nuzhat</a> ·
    <a href="{{ url_for('findings') }}">Review Findings</a> ·
    <a href="{{ url_for('modern_index') }}" style="color:#a78bfa;font-weight:600;">✨ Modern Course (new)</a>
  </nav>
</header>
{{ body|safe }}
<footer>Local preview · generated from the SQLite content store · courses/accounting-track</footer>
</body></html>
"""


def render(title, body_html):
    return render_template_string(BASE, title=title, body=body_html)


@app.route("/")
def home():
    con = db()
    course = con.execute("SELECT * FROM course LIMIT 1").fetchone()
    modules = con.execute("SELECT * FROM module ORDER BY position").fetchall()
    counts = con.execute(
        "SELECT severity, fix_status, COUNT(*) n FROM review_finding GROUP BY severity, fix_status"
    ).fetchall()
    open_blockers = con.execute(
        "SELECT COUNT(*) FROM review_finding WHERE fix_status='open' AND severity IN ('blocker','error')"
    ).fetchone()[0]
    fixed = con.execute("SELECT COUNT(*) FROM review_finding WHERE fix_status='fixed_in_repo_copy'").fetchone()[0]
    total = con.execute("SELECT COUNT(*) FROM review_finding").fetchone()[0]
    con.close()

    cards = ""
    for m in modules:
        color = STATUS_COLOR.get(m["status"], "#6b7280")
        day = f'<span class="pill">{m["day_mapping"]}</span>' if m["day_mapping"] else ""
        cards += f"""
        <a class="card" href="{url_for('module_view', slug=m['slug'])}">
          <span class="badge" style="background:{color}">{m['status'].replace('_',' ')}</span>
          <h3>{m['position']:02d} · {m['title']}</h3>
          <p>{m['summary'] or ''}</p>
          <div style="margin-top:8px;">{day}</div>
        </a>"""

    body = f"""
    <h1>{course['title']}</h1>
    <p class="tagline">{course['tagline']} — {course['audience']}</p>

    <div class="banner {'danger' if open_blockers else 'warn'}">
      <strong>{open_blockers} blocker/error finding(s) still open</strong> for Nuzhat &nbsp;·&nbsp;
      {fixed} of {total} total review findings already fixed in this repo copy &nbsp;·&nbsp;
      <a href="{url_for('findings')}">see full list →</a>
    </div>

    <h2>Modules ({len(modules)})</h2>
    <div class="grid">{cards}</div>
    """
    return render(course["title"], body)


@app.route("/module/<slug>")
def module_view(slug):
    con = db()
    m = con.execute("SELECT * FROM module WHERE slug=?", (slug,)).fetchone()
    if not m:
        abort(404)
    lessons = con.execute("SELECT * FROM lesson WHERE module_id=? ORDER BY position", (m["id"],)).fetchall()
    assignments = con.execute("SELECT * FROM assignment WHERE module_id=?", (m["id"],)).fetchall()
    resources = con.execute("SELECT * FROM resource WHERE module_id=?", (m["id"],)).fetchall()
    findings = con.execute("SELECT * FROM review_finding WHERE module_id=?", (m["id"],)).fetchall()
    objectives = con.execute("SELECT * FROM objective WHERE module_id=?", (m["id"],)).fetchall()
    con.close()

    color = STATUS_COLOR.get(m["status"], "#6b7280")
    body = f"""
    <nav class="crumbs"><a href="{url_for('home')}">Dashboard</a> / {m['title']}</nav>
    <span class="badge" style="background:{color}">{m['status'].replace('_',' ')}</span>
    <h1>{m['position']:02d} · {m['title']}</h1>
    <p class="tagline">{m['summary'] or ''}{' · maps to ' + m['day_mapping'] if m['day_mapping'] else ''}</p>
    """

    if objectives:
        body += "<h2>Learning objectives</h2><ul>"
        for o in objectives:
            tag = f' <span class="pill">MVR #{o["mvr_item"]}</span>' if o["mvr_item"] else ""
            body += f"<li>{o['text']}{tag}</li>"
        body += "</ul>"

    if lessons:
        body += "<h2>Lessons</h2><table><tr><th>#</th><th>Title</th><th>Session</th><th>Summary</th></tr>"
        for l in lessons:
            body += f"<tr><td>{l['position']}</td><td>{l['title']}</td><td>{l['session'] or '—'}</td><td>{l['summary'] or ''}</td></tr>"
        body += "</table>"

    if assignments:
        body += "<h2>Assignments</h2>"
        for a in assignments:
            grade = "Graded" if a["graded"] else "Practice"
            body += f"""<div class="card" style="margin-bottom:10px;">
              <h3>{a['title']} <span class="pill">{grade}</span> <span class="pill">~{a['est_hours']}h</span></h3>
              <p><strong>Instructions:</strong> {a['instructions']}</p>
              <p class="muted"><strong>Deliverable:</strong> {a['deliverable']}</p>
            </div>"""

    if resources:
        body += "<h2>Files</h2>"
        for r in resources:
            body += f"""<div class="resource-row">
              <div><strong>{r['title']}</strong> <span class="pill">{r['kind']}</span> <span class="pill">{r['role']}</span>
                <div class="muted" style="font-size:.82rem;">{r['notes'] or ''}</div></div>
              <div class="actions">
                <a href="{url_for('view_resource', relpath=r['file_path'])}">Preview →</a>
                <a href="{url_for('download_resource', relpath=r['file_path'])}">Download</a>
              </div>
            </div>"""

    if findings:
        body += "<h2>Review findings for this module</h2><ul class='findings'>"
        for f in findings:
            sc = SEV_COLOR.get(f["severity"], "#888")
            body += (f"<li><span class='sev' style='color:{sc}'>{f['severity']}</span> "
                      f"<span class='status-{f['fix_status']}'>[{f['fix_status'].replace('_',' ')}]</span> "
                      f"— {f['finding']}</li>")
        body += "</ul>"

    return render(m["title"], body)


@app.route("/findings")
def findings():
    con = db()
    rows = con.execute("""
        SELECT f.*, m.title as mtitle, m.slug as mslug FROM review_finding f
        JOIN module m ON m.id = f.module_id
        ORDER BY CASE f.severity WHEN 'blocker' THEN 0 WHEN 'error' THEN 1 WHEN 'improvement' THEN 2 ELSE 3 END, m.position
    """).fetchall()
    con.close()
    body = "<h1>Review Findings</h1><p class='tagline'>All issues found during the July 2026 review, ranked by severity.</p>"
    body += "<table><tr><th>Severity</th><th>Module</th><th>Finding</th><th>Status</th></tr>"
    for f in rows:
        sc = SEV_COLOR.get(f["severity"], "#888")
        body += (f"<tr><td class='sev' style='color:{sc}'>{f['severity']}</td>"
                  f"<td><a href='{url_for('module_view', slug=f['mslug'])}'>{f['mtitle']}</a></td>"
                  f"<td>{f['finding']}</td>"
                  f"<td class='status-{f['fix_status']}'>{f['fix_status'].replace('_',' ')}</td></tr>")
    body += "</table>"
    return render("Review Findings", body)


FEEDBACK_FILES = [
    ("FEEDBACK_FOR_NUZHAT.md", "Full Feedback Document"),
    ("JOB_MARKET_RESEARCH_2026.md", "Job Market Research (2026)"),
    ("NEXT_STEPS.md", "Next Steps"),
]


@app.route("/feedback")
def feedback_index():
    body = "<h1>Feedback for Nuzhat</h1><p class='tagline'>Isolated review deliverables — safe to hand off directly.</p><div class='grid'>"
    for fname, label in FEEDBACK_FILES:
        body += f"""<a class="card" href="{url_for('feedback_doc', fname=fname)}">
          <h3>{label}</h3><p>feedback-for-nuzhat/{fname}</p></a>"""
    body += f"""<a class="card" href="{url_for('view_resource', relpath='shared/MASTER_COMPANY_PROFILE.md')}">
      <h3>Master Company Profile</h3><p>shared/MASTER_COMPANY_PROFILE.md</p></a>"""
    body += "</div>"
    return render("Feedback for Nuzhat", body)


@app.route("/feedback/<fname>")
def feedback_doc(fname):
    if fname not in [f for f, _ in FEEDBACK_FILES]:
        abort(404)
    path = os.path.join(TRACK_ROOT, "feedback-for-nuzhat", fname)
    text = open(path, encoding="utf-8").read()
    html = md.markdown(text, extensions=["tables", "fenced_code"])
    body = f"<nav class='crumbs'><a href='{url_for('feedback_index')}'>Feedback</a> / {fname}</nav><div class='markdown-body'>{html}</div>"
    return render(fname, body)


def _safe_path(relpath):
    full = os.path.normpath(os.path.join(TRACK_ROOT, relpath))
    if not full.startswith(TRACK_ROOT):
        abort(403)
    if not os.path.exists(full):
        abort(404)
    return full


@app.route("/download/<path:relpath>")
def download_resource(relpath):
    full = _safe_path(relpath)
    if os.path.isdir(full):
        abort(400)
    return send_file(full, as_attachment=True)


@app.route("/view/<path:relpath>")
def view_resource(relpath):
    full = _safe_path(relpath)
    ext = os.path.splitext(full)[1].lower()
    crumb = f"<nav class='crumbs'><a href='{url_for('home')}'>Dashboard</a> / {relpath}</nav>"
    dl = f"<p><a href='{url_for('download_resource', relpath=relpath)}'>⬇ Download original file</a></p>"

    if os.path.isdir(full):
        files = sorted(os.listdir(full))
        body = crumb + f"<h1>{os.path.basename(full)}/</h1><ul>"
        for fn in files:
            rp = os.path.join(relpath, fn).replace("\\", "/")
            body += f"<li><a href='{url_for('view_resource', relpath=rp)}'>{fn}</a></li>"
        body += "</ul>"
        return render(relpath, body)

    if ext == ".md":
        text = open(full, encoding="utf-8").read()
        html = md.markdown(text, extensions=["tables", "fenced_code"])
        return render(relpath, crumb + f"<div class='markdown-body'>{html}</div>")

    if ext == ".docx":
        def convert_image(image):
            with image.open() as img_bytes:
                b64 = base64.b64encode(img_bytes.read()).decode("ascii")
            return {"src": f"data:{image.content_type};base64,{b64}"}

        with open(full, "rb") as f:
            result = mammoth.convert_to_html(
                f, convert_image=mammoth.images.img_element(convert_image)
            )
        warn_html = ""
        banner = ""
        if "_drafts-needs-redo" in full or "kaizen" in full.lower():
            banner = "<div class='banner danger'><strong>⚠ This draft fails the brief</strong> — wrong company (kaizen, BDT), see the feedback doc for the full writeup and redo instructions.</div>"
        body = crumb + banner + f"<h1>{os.path.basename(full)}</h1>" + dl + f"<div class='doc-preview'>{result.value}</div>"
        return render(os.path.basename(full), body)

    if ext == ".xlsx":
        wb = openpyxl.load_workbook(full, data_only=True)
        wb_f = openpyxl.load_workbook(full, data_only=False)
        tabs = "".join(
            f"<a href='#{s}' onclick=\"document.querySelectorAll('.xlsx-sheet').forEach(e=>e.style.display='none');document.getElementById('sheet-{i}').style.display='block';document.querySelectorAll('.tabbar a').forEach(a=>a.classList.remove('active'));this.classList.add('active');return false;\">{s}</a>"
            for i, s in enumerate(wb.sheetnames)
        )
        banner = ""
        if "kaizen" in full.lower():
            banner = "<div class='banner danger'><strong>⚠ Wrong company / wrong currency</strong> — this export is from an unrelated 'kaizen' QBO demo company in BDT, not Skarion Manufacturing in USD. See feedback for the redo plan.</div>"

        # Detect whether any formula cell is missing its cached value (happens to files this
        # project re-saved with openpyxl); if so, recalculate once with the `formulas` engine.
        needs_recalc = False
        for name in wb.sheetnames:
            ws_f, ws_v = wb_f[name], wb[name]
            for row_f, row_v in zip(ws_f.iter_rows(max_row=min(ws_f.max_row, 60)), ws_v.iter_rows(max_row=min(ws_v.max_row, 60))):
                for cf, cv in zip(row_f, row_v):
                    if isinstance(cf.value, str) and cf.value.startswith("=") and cv.value is None:
                        needs_recalc = True
                        break
                if needs_recalc:
                    break
            if needs_recalc:
                break
        computed = get_computed_values(full) if needs_recalc else {}
        recalced_note = ""
        if needs_recalc and "__error__" not in computed:
            recalced_note = "<div class='banner warn'>ℹ Some formula results were recalculated for this preview (openpyxl doesn't cache formula output on save) — values shown are correct, and the original formulas are untouched.</div>"
        elif needs_recalc:
            recalced_note = f"<div class='banner warn'>⚠ Could not recalculate formulas for preview ({computed.get('__error__','unknown error')[:120]}) — blank cells below are formula cells; open the downloaded file in Excel to see live values.</div>"

        sheets_html = ""
        MAX_ROWS, MAX_COLS = 80, 20
        for i, name in enumerate(wb.sheetnames):
            ws = wb[name]
            ws_f = wb_f[name]
            sheet_key = name.upper()
            display = "block" if i == 0 else "none"
            sheets_html += f"<div class='xlsx-sheet' id='sheet-{i}' style='display:{display}'><h4>{name}</h4><table>"
            for row, row_f in zip(
                ws.iter_rows(min_row=1, max_row=min(ws.max_row, MAX_ROWS), max_col=min(ws.max_column, MAX_COLS)),
                ws_f.iter_rows(min_row=1, max_row=min(ws.max_row, MAX_ROWS), max_col=min(ws.max_column, MAX_COLS)),
            ):
                sheets_html += "<tr>"
                for c, cf in zip(row, row_f):
                    val = c.value
                    if val is None and isinstance(cf.value, str) and cf.value.startswith("="):
                        val = computed.get(sheet_key, {}).get(c.coordinate)
                    val = "" if val is None else (f"{val:,.2f}" if isinstance(val, float) else str(val))
                    sheets_html += f"<td>{val}</td>"
                sheets_html += "</tr>"
            sheets_html += "</table>"
            if ws.max_row > MAX_ROWS or ws.max_column > MAX_COLS:
                sheets_html += f"<p class='muted'>Showing first {MAX_ROWS} rows / {MAX_COLS} cols of {ws.max_row}x{ws.max_column} — download for full data.</p>"
            sheets_html += "</div>"
        body = crumb + banner + recalced_note + f"<h1>{os.path.basename(full)}</h1>" + dl + f"<div class='tabbar'>{tabs}</div>{sheets_html}"
        return render(os.path.basename(full), body)

    if ext == ".pdf":
        rp = relpath.replace("\\", "/")
        body = crumb + f"<h1>{os.path.basename(full)}</h1>" + dl + f"<iframe src='{url_for('raw_file', relpath=rp)}' style='width:100%;height:85vh;border:1px solid #262b33;border-radius:8px;'></iframe>"
        return render(os.path.basename(full), body)

    # fallback: offer download only
    body = crumb + f"<h1>{os.path.basename(full)}</h1><p>No inline preview for this file type.</p>" + dl
    return render(os.path.basename(full), body)


@app.route("/raw/<path:relpath>")
def raw_file(relpath):
    full = _safe_path(relpath)
    return send_file(full)


# ---------- Modern interactive course (claude-made-course) ----------

MODULE_META = [
    ("00-course-design", "00", "Course Design & Backbone"),
    ("01-job-market-research", "01", "Job Market Research"),
    ("02-quickbooks-online-lab", "02", "QuickBooks Online Lab"),
    ("03-core-accounting-practice", "03", "Core Accounting Practice"),
    ("04-bank-reconciliation-lab", "04", "Bank Reconciliation Lab"),
    ("05-accounts-payable-simulation", "05", "Accounts Payable Simulation"),
    ("06-interview-preparation", "06", "Interview Preparation"),
    ("07-excel-toolkit", "07", "Excel Toolkit"),
    ("08-payroll-fundamentals", "08", "Payroll Fundamentals"),
    ("09-month-end-close", "09", "Month-End Close"),
]


@app.route("/modern")
def modern_index():
    cards = ""
    for slug, num, title in MODULE_META:
        html_path = os.path.join(MODERN_MODULES, slug, "index.html")
        ready = os.path.exists(html_path)
        badge = "<span class='badge' style='background:#7c5cff'>ready</span>" if ready else "<span class='badge' style='background:#6b7280'>coming soon</span>"
        href = url_for("modern_module", slug=slug) if ready else "#"
        cards += f"""<a class="card" href="{href}" style="{'opacity:.5;pointer-events:none;' if not ready else ''}">
          {badge}<h3>{num} · {title}</h3><p>Interactive micro-lesson + quiz + SCORM export</p></a>"""
    body = f"""
    <h1>✨ Modern Interactive Course</h1>
    <p class="tagline">Chunked, interactive lessons built on the fixed course content — each one exports as a SCORM 1.2 package for LMS import.</p>
    <div class="grid">{cards}</div>
    """
    return render("Modern Course", body)


@app.route("/modern/<slug>")
def modern_module(slug):
    html_path = os.path.join(MODERN_MODULES, slug, "index.html")
    if not os.path.exists(html_path):
        abort(404)
    return send_file(html_path)


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5057, debug=False)
